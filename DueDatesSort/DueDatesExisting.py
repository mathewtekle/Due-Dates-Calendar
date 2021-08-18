import openpyxl
import os
from openpyxl import load_workbook, Workbook
from openpyxl import worksheet
from openpyxl import reader
from openpyxl.styles import Font
from DueDateWriterOnly import CreateWorkbook

class Sheets(CreateWorkbook):
    if os.path.exists('dueDates2021.xlsx'):  # if excel file exists, load the workbook and sheets
        wb = load_workbook('dueDates2021.xlsx')
        sheet = wb.active
        make_book = CreateWorkbook(wb, sheet)
        if 'Homework' != sheet['A1']:  # if this header is not there, can assume formatting is off
            make_book.sheet_formatting(sheet)
        make_book.add_info_sheet(sheet)
        wb.save('dueDates2021.xlsx')
    else:  # if excel file does not exist, create one.
        wb = Workbook()
        wb.create_sheet('Calendar')
        sheet = wb.active
        make_new_book = CreateWorkbook(wb, sheet)
        make_new_book.sheet_formatting(sheet)
        make_new_book.add_info_sheet(sheet)
        wb.save('dueDates2021.xlsx')
